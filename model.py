import gurobipy as gp
from gurobipy import GRB
import numpy as np
import pandas as pd
import os
import openpyxl
from datetime import datetime, timedelta
import locale


folder_path = "Data"  # Make sure the folder path is correct

# Read each CSV file into a DataFrame
df_price = pd.read_csv(os.path.join(folder_path, "Prices.csv"))
df_chargesvar = pd.read_csv(os.path.join(folder_path, "Charges_var.csv"))
df_prod = pd.read_csv(os.path.join(folder_path, "Production.csv"))
df_plantation = pd.read_csv(os.path.join(folder_path, "plantation.csv"))
df_month_index = pd.read_csv(os.path.join(folder_path, "month_index.csv"))
df_sim=pd.read_csv(os.path.join(folder_path, "Simulation.csv"))
# Now you can work with these DataFrames as needed
df_price = df_price.fillna(0)
df_prod=df_prod.fillna(0)
serre_sau_dict = df_sim.set_index('serre')['SAU(ha)'].to_dict()

scenario_variety_mapping = df_prod[['Scénario', 'variété 23-24']]
sum_last_37 = df_prod.iloc[:, -37:].sum(axis=1)

# Create a dictionary to associate scenario with the sum of the last 37 columns
scenario_prod = dict(zip(df_prod['Scénario'], sum_last_37))

# Initialize an empty dictionary to store the mappings
variety_scenario_dict = {}

# Iterate through each row of the DataFrame
for index, row in scenario_variety_mapping.iterrows():
    scenario = row['Scénario']
    variety = row['variété 23-24']
    
    # If the variety is not already in the dictionary, add it with an empty list
    if variety not in variety_scenario_dict:
        variety_scenario_dict[variety] = []
    
    # Append the scenario to the list of scenarios for this variety
    variety_scenario_dict[variety].append(scenario)


locale.setlocale(locale.LC_TIME, 'fr_FR.UTF-8')

# Define the start date
start_date = datetime(2024, 1, 1)

# Function to get the month index
def get_month_index(month):
    month_dict = {
        'Janvier': 1,
        'Février': 2,
        'Mars': 3,
        'Avril': 4,
        'Mai': 5,
        'Juin': 6,
        'Juillet': 7,
        'Aout': 8,
        'Septembre': 9,
        'Octobre': 10,
        'Novembre': 11,
        'Décembre': 12
    }
    return month_dict[month]

# Generate the data
data = []
months = [
    'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Aout', 'Septembre', 'Octobre', 'Novembre', 'Décembre'
]

num_weeks=84

for index in range(1,num_weeks+1):
    month_index = start_date.month
    month_name = months[month_index - 1]  # Adjust for zero-based indexing

# Capitalize the first letter of the month name
    month = month_name.capitalize()

    # month = start_date.strftime('%B').capitalize()
    month_index = get_month_index(month)
    data.append([month, start_date.strftime('%d/%m/%Y'), month_index, index])
    start_date += timedelta(weeks=1)

# Create DataFrame
df_week_index = pd.DataFrame(data, columns=['Month', 'Date', 'Month Index', 'Week Index'])

# Save to CSV
month_to_week_indices = {}

# Iterate through the DataFrame to populate the dictionary
for index, row in df_week_index.iterrows():
    month = row['Month']
    week_index = row['Week Index']
    
    if month not in month_to_week_indices:
        month_to_week_indices[month] = []
    
    month_to_week_indices[month].append(week_index)





scenario_delai_dict = {}
scenario_duree_dict = {}
scenario_mois_dict = {}
scenario_culture={}
scenario_cmo={}
scenario_vitesse={}
scenario_cv={}
for index,row in df_chargesvar.iterrows():
    scenario = row['Scénario']
    cmo=row["Coûts de Main d'œuvre par kg de production"]
    vitesse=row["Vitesse de main d'œuvre kg/personne/jour"]
    cv=row["Coûts variable par hectar"]
    scenario_cmo[scenario]=cmo
    scenario_vitesse[scenario]=vitesse
    scenario_cv[scenario]=cv
for index, row in df_prod.iterrows():
    scenario = row['Scénario']
    delai = row['Délai pour début de production']
    duree = row['Durée de production en semaine']
    mois = row['Mois']
    culture=row['Culture']
    
    scenario_delai_dict[scenario] = delai
    scenario_duree_dict[scenario] = duree
    scenario_mois_dict[scenario] = mois
    scenario_culture[scenario] =culture
    
scenarios=list(scenario_culture.keys())
price={}
price["Framboise"]=np.array(df_price.iloc[0,1:])[:-1]
price["Mure"]=np.array(df_price.iloc[1,1:])[:-1]
secteur_serre_dict = {}

# Iterate over the rows of the DataFrame
for index, row in df_sim.iterrows():
    secteur = row['secteur']
    serre = row['serre']
    
    # If the secteur is not in the dictionary yet, create a new list
    # Otherwise, append the serre to the existing list
    if secteur not in secteur_serre_dict:
        secteur_serre_dict[int(secteur)] = [int(serre)]
    else:
        secteur_serre_dict[int(secteur)].append(int(serre))
serre_secteur_dict = {serre: secteur for secteur, serres in secteur_serre_dict.items() for serre in serres}
num_serre=df_sim.shape[0]
num_sect=len(secteur_serre_dict)



prod_mat = np.matrix(df_prod.iloc[:, 8:])

prod = {}

def padded_dot(a, b):
    # Get the lengths of the arrays
    len_a = a.shape[1]
    len_b = b.shape[1]

    # Determine the maximum length
    max_len = max(len_a, len_b)

    # Pad arrays with zeros only to the right side
    a_padded = np.pad(a, ((0, 0), (0, max_len - len_a)), mode='constant', constant_values=0)
    b_padded = np.pad(b, ((0, 0), (0, max_len - len_b)), mode='constant', constant_values=0)

    # Perform dot product
    result = np.dot(a_padded, b_padded.T)

    return result


for i in range(num_serre):
    for j in scenarios:

        for t in month_to_week_indices[scenario_mois_dict[j]]:
            
            if j in variety_scenario_dict["Adelita"]:
                    price_array = np.array(price[scenario_culture[j]][t-1 + scenario_delai_dict[j] :])+5
                    prod_mat_array = np.array(prod_mat[scenarios.index(j), :])
                    prod[(i, j, t)] = serre_sau_dict[i + 1] *padded_dot(
                    price_array.reshape(1, -1),
                    prod_mat_array.reshape(1, -1)
                )
                    
            else:
                    
                    price_array = np.array(price[scenario_culture[j]][t-1 + scenario_delai_dict[j]:])
                    prod_mat_array = np.array(prod_mat[scenarios.index(j), :])
                    prod[(i, j, t)] = serre_sau_dict[i + 1] *padded_dot(
                    price_array.reshape(1, -1),
                    prod_mat_array.reshape(1, -1)
                )
                    
try:
    # Create a new model
    m = gp.Model("portfolio")

    
    choices={}
    for i in range(num_serre):
        for j in scenarios:
            for t in month_to_week_indices[scenario_mois_dict[j]]:
                choices[(i,j,t)]= m.addVar(vtype=GRB.BINARY, name=f'choice_{i}_{j}_{t}')
    
   
    for i in range(num_serre):
        m.addConstr(gp.quicksum(choices[(i, j, t)] for j in scenarios for t in month_to_week_indices[scenario_mois_dict[j]]) == 1, f"constraint_{i}")
   
    # Meme scenario par secteur
    for i in range(num_sect):
        ref = secteur_serre_dict[i+1][0]
        # m.addConstr(gp.quicksum(choix_semaines[(i, t)] for t in range(90)))
        for j in secteur_serre_dict[i+1]:
            for k in scenarios:
                for t in month_to_week_indices[scenario_mois_dict[k]]:
                    m.addConstr(choices[(j-1,k,t)] 
                                == choices[(ref-1,k,t)], f'c_0_{j}_{k}_{t}')
                

    # Cette contrainte est obtenue par la contraposé de la cinquième contrainte
    for i in range(21):
        for t in month_to_week_indices[scenario_mois_dict[5]]:
            m.addConstr(choices[(i,5,t)]==0)

    #Contrainte limite de main d'oeuvre
    
    #contrainte choix
    for i in range(num_serre):
        if i!=19 and i!=20:
            for j in variety_scenario_dict["Clara"]:
                for t in month_to_week_indices[scenario_mois_dict[j]]:
                    m.addConstr(choices[(i,j,t)]==0)
            for j in variety_scenario_dict["LAURITA"]:
                for t in month_to_week_indices[scenario_mois_dict[j]]:
                    m.addConstr(choices[(i,j,t)]==0)
        else:
            if serre_sau_dict[i+1]>2.87:
                for j in variety_scenario_dict["Clara"]:
                    for t in month_to_week_indices[scenario_mois_dict[j]]:
                        m.addConstr(choices[(i,j,t)]==0)
                for j in variety_scenario_dict["LAURITA"]:
                    for t in month_to_week_indices[scenario_mois_dict[j]]:
                        m.addConstr(choices[(i,j,t)]==0)
           

    for s in range(1,91):
        scenarios_time_dict={}
        for j in scenarios:
            scenarios_time_dict[j]=[]
            if j in [4,5,20]:
                for t in month_to_week_indices[scenario_mois_dict[j]]:
                    if t+scenario_delai_dict[j]<s and s<t+scenario_delai_dict[j]+38:
                        scenarios_time_dict[j].append(t)
                    else: 
                        pass

            else:

                for t in month_to_week_indices[scenario_mois_dict[j]]:
                    if t+scenario_delai_dict[j]<s and s<t+scenario_delai_dict[j]+scenario_duree_dict[j]+1:
                        scenarios_time_dict[j].append(t)
                    else: 
                        pass
        m.addConstr(gp.quicksum(choices[(i,k,t)]*(serre_sau_dict[i+1]/scenario_vitesse[k])
            *prod_mat[scenarios.index(k),s-t-scenario_delai_dict[k]-1]
            for k in scenarios 
            for t in scenarios_time_dict[k] if scenarios_time_dict[k]!=[]
            
                for i in range(num_serre)) <= 7*600, f'mo_{s}')





    CA_expr = gp.quicksum( choices[(i,j,t)]* 
                          prod[(i, j, t)] 
                           for i in range(num_serre)
                     for j in scenarios for t in month_to_week_indices[scenario_mois_dict[j]]
                        )
    CMO_expr = gp.quicksum(choices[(i,j,t)] * scenario_prod[j]
                        #    can be changed 
                           * serre_sau_dict[i+1] * scenario_cmo[j] 
                       for j in scenarios for t in month_to_week_indices[scenario_mois_dict[j]]
                        for i in range(num_serre))
    CV_expr = gp.quicksum(choices[(i,j,t)]* 
                          serre_sau_dict[i+1] * 
                          scenario_cv[j] for j in scenarios for t in month_to_week_indices[scenario_mois_dict[j]]
                        for i in range(num_serre))

    # Set objective
    m.update()
    m.setObjective((CA_expr - CV_expr- CMO_expr), GRB.MAXIMIZE)

    # Optimize model
    m.optimize()

    # for v in m.getVars():
    #     print(f"{v.VarName} {v.X:g}")
    # After the optimization is done, you can print the expressions
    print("Value of CA_expr:", CA_expr.getValue())
    print("Value of CMO_expr:", CMO_expr.getValue())
    print("Value of CV_expr:", CV_expr.getValue())


    print(f"Obj: {m.ObjVal:g}")
    


except gp.GurobiError as e:
    print(f"Error code {e.errno}: {e}")

except AttributeError:
    print("Encountered an attribute error")


scenario_chosen, semaines_chosen=[],[]
if m.status == GRB.Status.OPTIMAL:
        # Get selected variables with value == 1
        list_var = [v.varName for v in m.getVars() if v.x == 1]

for i in list_var:
    scenario_chosen.append(int(i.split("_")[2]))
    semaines_chosen.append(int(i.split("_")[3]))

def write_arrays_to_excel(filename, scenario_chosen, semaines_chosen):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    # Select the worksheet named "Simulation"
    ws = wb["Simulation"]
    
    # Write scenario_chosen to column E starting from cell E23
    for i, value in enumerate(scenario_chosen):
        ws.cell(row=23+i, column=5, value=value)
    
    # Write semaines_chosen to column M starting from cell M23
    for i, value in enumerate(semaines_chosen):
        ws.cell(row=23+i, column=13, value=value)
    
    # Save the workbook
    wb.save(filename)
    print("All safe and sound")

# Example usage:

write_arrays_to_excel("Modèle_Belfaa_updated_linearized_1303_simul_noQ.xlsx", scenario_chosen, semaines_chosen)
